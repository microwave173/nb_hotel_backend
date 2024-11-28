from openpyxl import load_workbook


workbook = load_workbook(filename="in.xlsx")
sheet = workbook["Sheet1"]

workbook1 = load_workbook(filename="out.xlsx")
sheet1 = workbook1["Sheet1"]  # 4, 7


def get_data():
    ret = []
    for i in range(1, 27):
        row = []
        for col in ['A', 'B', 'C', 'D', 'E']:
            temp = sheet[col + str(i)].value
            if temp is None:
                row.append([''])
            else:
                row.append(str(temp).split('，'))
        ret.append(row)
    return ret


def dict2xlsx(data):
    i1 = 0
    for i in range(4, 30):
        j1 = 0
        for j in range(7, 27, 4):
            sheet1.cell(row=i, column=j, value=data[i1][j1]['curTemperature'])
            sheet1.cell(row=i, column=j+1, value=data[i1][j1]['tarTemperature'])
            sheet1.cell(row=i, column=j+2, value=data[i1][j1]['acMode'])
            sheet1.cell(row=i, column=j+3, value=data[i1][j1]['du'])
            j1 += 1
        i1 += 1
    workbook1.save('out.xlsx')
