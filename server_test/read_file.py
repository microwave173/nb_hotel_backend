from openpyxl import load_workbook

folder = 'heat'

workbook = load_workbook(filename=f"{folder}/in.xlsx")
sheet = workbook["Sheet1"]

workbook1 = load_workbook(filename=f"{folder}/out.xlsx")
sheet1 = workbook1["Sheet1"]  # 4, 7


def get_data():
    ret = []
    for i in range(1, 27):
        row = []
        for col in ['A', 'B', 'C', 'D', 'E']:
            temp = sheet[col + str(i)].value
            if temp is None:
                row.append([''])
            else:
                row.append(str(temp).split('，'))
        ret.append(row)
    return ret


def dict2xlsx(data, logs, queues, cost_list):
    i1 = 0
    for i in range(4, 30):
        j1 = 0
        for j in range(7, 27, 4):
            sheet1.cell(row=i, column=j, value=data[i1][j1]['curTemperature'])
            sheet1.cell(row=i, column=j+1, value=data[i1][j1]['tarTemperature'])
            sheet1.cell(row=i, column=j+2, value=data[i1][j1]['acMode'])
            sheet1.cell(row=i, column=j+3, value=data[i1][j1]['cost'])
            j1 += 1
        i1 += 1

    i1 = 0
    for i in range(34, 34 + len(logs[0])):
        j1 = 0
        for j in range(2, 4*5+2, 4):
            if i1 < len(logs[j1]):
                sheet1.cell(row=i, column=j, value=logs[j1][i1]['logType'])
                sheet1.cell(row=i, column=j+1, value=logs[j1][i1]['costRate'])
                sheet1.cell(row=i, column=j+2, value=logs[j1][i1]['costSum'])
                sheet1.cell(row=i, column=j+3, value=logs[j1][i1]['logTime'])
            j1 += 1
        i1 += 1

    i1 = 0
    for i in range(4, 30):
        for j in range(28, 28 + 13):
            sheet1.cell(row=i, column=j, value='')
        for j in range(len(queues[i1]['serve_queue'])):
            sheet1.cell(row=i, column=j+28, value=queues[i1]['serve_queue'][j])
        for j in range(len(queues[i1]['wait_queue'])):
            sheet1.cell(row=i, column=j+31, value=queues[i1]['wait_queue'][j])
        i1 += 1

    for j in range(2, 7):
        sheet1.cell(row=30, column=j, value=cost_list[j - 2])

    workbook1.save(f'{folder}/out.xlsx')
